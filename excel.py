import logging
import sys

import openpyxl

from models import Category

"""
Class create excel file with required headers as soon as it initializes 
it will write all the scrape data into excel file
"""


class Excel:
    def __init__(self, output_excel_file: str = 'output.xlsx'):
        self.op_excel_file = output_excel_file
        self.setup_op_file_with_header()

    def write_excel_file_by_row(self, category_obj: Category) -> None:
        try:
            op_workbook = openpyxl.load_workbook(filename=self.op_excel_file)
            worksheet = op_workbook.active

            max_row = worksheet.max_row + 1

            worksheet.cell(row=max_row, column=1).value = category_obj.category_name
            worksheet.cell(row=max_row, column=2).value = category_obj.course.course_name
            worksheet.cell(row=max_row, column=3).value = category_obj.course.description
            worksheet.cell(row=max_row, column=4).value = '\n'.join(category_obj.course.course_features)
            worksheet.cell(row=max_row, column=5).value = category_obj.course.price
            worksheet.cell(row=max_row, column=6).value = '\n'.join(category_obj.course.what_youll_learn)

            worksheet.cell(row=max_row, column=7).value = '\n'.join(category_obj.course.timings)
            worksheet.cell(row=max_row, column=8).value = '\n'.join(category_obj.course.requirements)
            worksheet.cell(row=max_row, column=9).value = '\n'.join(category_obj.course.curriculum)
            worksheet.cell(row=max_row, column=10).value = '\n'.join(category_obj.course.mentor_names)

            op_workbook.save(filename=self.op_excel_file)
        finally:
            op_workbook.save(filename=self.op_excel_file)
            op_workbook.close()

    def setup_op_file_with_header(self) -> None:
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            headers = ['Category', 'Course Name', 'Description', 'Features', 'Price', "What You'll Learn", 'Timings',
                       'Requirements',
                       'Curriculum', 'Mentor Names']

            worksheet.append(headers)
            for cell in worksheet[f'1:1']:  # first row
                cell.font = openpyxl.styles.Font(bold=True)

            workbook.save(self.op_excel_file)
        except PermissionError:
            print('[ERROR] file is opened please close the file first')
            sys.exit()


if __name__ == '__main__':
    exel = Excel(output_excel_file='output.xlsx')
    exel.setup_op_file_with_header()
